"""Build the budget line sheets with full ancestry.

Reads the converted .xlsx files in 02_cleaned/ and writes two workbooks to 03_extracted/:
    SituationChap_LIGNES_BUDGETAIRES_TOTAL.xlsx    -> every leaf line, all years
    SituationChap_LIGNES_BUDGETAIRES_COMMUNS.xlsx  -> only lines present in every year

Output columns (one row per leaf Ligne, all years stacked):
    Year, Chap, Chap_Intitule, Prog, Prog_Intitule, Reg, Reg_Intitule,
    Proj, Proj_Intitule, Lb, Intitule,
    Credits_Ouverts_Vises, Virements_En_Plus_Vises, Virements_En_Moins_Vises,
    Total_Engage_Vises

Raw layout (per year): each budget line spans 2 physical rows
  - top row    : Intitulé (H), amounts (K, M, O, AH on the "Visés" sub-column)
  - bottom row : Chap (B), Prog (C), Reg (D), Proj (E), Lb (G)
Only ONE of B/C/D/E/G is filled per pair = that row's level.
"""
from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent
RAW_DIR = DATA_DIR / "02_cleaned"
OUT_DIR = DATA_DIR / "03_extracted"
OUT_DIR.mkdir(exist_ok=True)

# Hierarchy code columns (bottom row of each pair)
COL_CHAP, COL_PROG, COL_REG, COL_PROJ, COL_LB = "B", "C", "D", "E", "G"

# Top-row columns (Visés sub-column)
COL_INTITULE                = "H"
COL_CREDITS_OUVERTS_VISES   = "K"
COL_VIREMENTS_EN_PLUS_VISES = "M"
COL_VIREMENTS_EN_MOINS_VISES = "O"
COL_TOTAL_ENGAGE_VISES      = "AH"

DATA_START_ROW = 6  # rows 4-5 are headers

# Manager said these chapters are not relevant
EXCLUDED_CHAPTERS = {"3200106002", "3200106001"}

LEVELS = ["Chap", "Prog", "Reg", "Proj", "Lb"]
CODE_COLS = [COL_CHAP, COL_PROG, COL_REG, COL_PROJ, COL_LB]


def _txt(ws, col: str, row: int):
    v = ws[f"{col}{row}"].value
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def _code(ws, col: str, row: int):
    """Hierarchy code as string, preserving leading zeros."""
    v = ws[f"{col}{row}"].value
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


def _num(ws, col: str, row: int):
    v = ws[f"{col}{row}"].value
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _extract_year(path: Path) -> int | None:
    m = re.search(r"(\d{4})", path.stem)
    return int(m.group(1)) if m else None


def flatten_one(xlsx_path: Path) -> list[dict]:
    """Return one dict per leaf Ligne for this year's file."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    year = _extract_year(xlsx_path)

    # Running ancestry: code + intitulé at each level (0=Chap .. 4=Lb)
    codes  = [None] * 5
    labels = [None] * 5
    out: list[dict] = []

    r = DATA_START_ROW
    while r <= ws.max_row:
        # Find which level this pair belongs to
        depth = None  # 1..5
        code = None
        for i, col in enumerate(CODE_COLS, start=1):
            c = _code(ws, col, r + 1)
            if c is not None:
                depth = i
                code = c
                break

        intitule = _txt(ws, COL_INTITULE, r)

        if depth is None and intitule is None:
            r += 2
            continue  # blank pair

        if depth is not None:
            # Update ancestry: set this level, clear everything deeper
            codes[depth - 1]  = code
            labels[depth - 1] = intitule
            for j in range(depth, 5):
                codes[j]  = None
                labels[j] = None

            if depth == 5:  # leaf Ligne -> emit a row
                if codes[0] in EXCLUDED_CHAPTERS:
                    r += 2
                    continue
                out.append({
                    "Year": year,
                    "Chap": codes[0], "Chap_Intitule": labels[0],
                    "Prog": codes[1], "Prog_Intitule": labels[1],
                    "Reg":  codes[2], "Reg_Intitule":  labels[2],
                    "Proj": codes[3], "Proj_Intitule": labels[3],
                    "Lb":   codes[4], "Intitule":      labels[4],
                    "Credits_Ouverts_Vises":    _num(ws, COL_CREDITS_OUVERTS_VISES, r),
                    "Virements_En_Plus_Vises":  _num(ws, COL_VIREMENTS_EN_PLUS_VISES, r),
                    "Virements_En_Moins_Vises": _num(ws, COL_VIREMENTS_EN_MOINS_VISES, r),
                    "Total_Engage_Vises":       _num(ws, COL_TOTAL_ENGAGE_VISES, r),
                })
        r += 2

    return out


def main():
    files = sorted(
        p for p in RAW_DIR.glob("SituationChap-*.xlsx")
        if re.fullmatch(r"SituationChap-\d{4}", p.stem)
    )
    if not files:
        raise SystemExit(f"No yearly SituationChap-YYYY.xlsx in {RAW_DIR}")

    all_rows: list[dict] = []
    for f in files:
        rows = flatten_one(f)
        print(f"{f.name}: {len(rows)} leaf lines")
        all_rows.extend(rows)

    df = pd.DataFrame(all_rows)
    cols = ["Year",
            "Chap", "Chap_Intitule",
            "Prog", "Prog_Intitule",
            "Reg",  "Reg_Intitule",
            "Proj", "Proj_Intitule",
            "Lb",   "Intitule",
            "Credits_Ouverts_Vises",
            "Virements_En_Plus_Vises",
            "Virements_En_Moins_Vises",
            "Total_Engage_Vises"]
    df = df[cols]

    # Stable identifier for each leaf line across years
    df["Line_Key"] = (
        df["Chap"].astype(str) + "-" +
        df["Prog"].astype(str) + "-" +
        df["Reg"].astype(str)  + "-" +
        df["Proj"].astype(str) + "-" +
        df["Lb"].astype(str)
    )

    out_path = OUT_DIR / "SituationChap_LIGNES_BUDGETAIRES_TOTAL.xlsx"
    df_total = df.copy()
    df_total["Year"] = df_total["Year"].astype(str)  # text so Power BI treats it as category
    df_total.to_excel(out_path, index=False)
    print(f"\n{len(df)} rows -> {out_path}")

    # Keep only lines whose full path appears in EVERY year
    n_years = df["Year"].nunique()
    years_per_key = df.groupby("Line_Key")["Year"].nunique()
    common_keys = years_per_key[years_per_key == n_years].index
    communs = df[df["Line_Key"].isin(common_keys)].drop(columns="Line_Key")

    out_communs = OUT_DIR / "SituationChap_LIGNES_BUDGETAIRES_COMMUNS.xlsx"
    communs.to_excel(out_communs, index=False)
    print(f"{len(communs)} rows ({len(common_keys)} lines x {n_years} years) -> {out_communs}")


if __name__ == "__main__":
    main()
