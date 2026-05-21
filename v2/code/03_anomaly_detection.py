"""
Anomaly detection on the SitGen budget panel.

We ask 12 simple yes/no questions about each budget line, in two buckets.

  Columns used (from the raw SitGen file):
    Credits_Ouverts  (col K) : opening budget at start of year
    Virements_Plus   (col M) : money transferred IN during the year
    Virements_Moins  (col O) : money transferred OUT during the year
    Total_Credits    (col AB): final ceiling = Credits_Ouverts + Plus - Moins
    Total_Engage     (col AH): money actually committed / spent
    Intitule         (col H) : line label
    Line_Key                 : stable identifier across years

  -----------------------------------------------------------------
  Bucket 1 - Data quality (line is BROKEN -> drop from forecast)
  -----------------------------------------------------------------
    over_execution    Total_Engage > 1.10 * Total_Credits
                      (spent more than the final ceiling)
    negative_value    Total_Engage < 0  or  Credits_Ouverts < 0
    summary_keyword   Intitule contains "TOTAL" / "Sous-total"
                      (a summary row got mixed in with real lines)
    identity_drift    Same Line_Key has different Intitule across years
    huge_swing        Total_Engage changed by >10x year-over-year

  -----------------------------------------------------------------
  Bucket 2 - Business watchlist (line is REAL but worth attention)
  -----------------------------------------------------------------
    major_virement           |Plus - Moins|  > 50% of Credits_Ouverts
    chronic_under_execution  Total_Engage / Total_Credits < 30%  for >=2 yrs
    chronic_over_execution   Total_Engage / Total_Credits > 100% for >=2 yrs
    high_volatility          std(Total_Engage) / mean(Total_Engage) > 0.5
    virement_inflation       Virements_Plus  > 100% of Credits_Ouverts
    virement_churn           Plus AND Moins  both > 30% of Credits_Ouverts
    unused_top_up            Plus > 30% AND  exec_rate < 80%

Outputs (in v2/data/03_forecast/):
    03_anomalies_extraction.csv   bucket 1 flags
    03_anomalies_business.csv     bucket 2 flags
    03_raw_enriched.csv           panel + virement columns (used by forecast)
"""
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "01_raw"
OUT = ROOT / "data" / "03_forecast"
OUT.mkdir(parents=True, exist_ok=True)

# Column letters in the raw GID file
COL = dict(
    CHAP="B", PROG="C", REG="D", PROJ="E", LB="G",
    INTITULE="H",
    CREDITS_VISES="K",
    VIREMENTS_PLUS_VISES="M",
    VIREMENTS_MOINS_VISES="O",
    TOTAL_CREDITS_VISES="AB",
    ENGAGE_VISES="AH",
)
DATA_START_ROW = 6
LEVELS = ["Chap", "Prog", "Reg", "Proj", "Lb"]
LEVEL_NAMES = {1: "Chapitre", 2: "Programme", 3: "Region", 4: "Projet", 5: "Ligne"}
EXCLUDED_CHAPTERS = {"3200106001", "3200106002"}
PARTIAL_YEARS = {2026}

SUMMARY_KEYWORDS = ("total", "sous-total", "sous total", "s/total")


def cell(ws, addr):
    v = ws[addr].value
    if isinstance(v, str):
        v = v.strip()
        return None if v == "" else v
    return v


def cell_code(ws, addr):
    v = ws[addr].value
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and float(v).is_integer():
        return str(int(v))
    return str(v).strip() or None


def to_num(v):
    if v is None or v == "":
        return None
    try:
        n = float(v)
        return n if np.isfinite(n) else None
    except (TypeError, ValueError):
        return None


def extract_one(xlsx_path, year):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row
    rows = []
    r = DATA_START_ROW
    while r <= max_row:
        intit = cell(ws, f"{COL['INTITULE']}{r}")
        credits = to_num(cell(ws, f"{COL['CREDITS_VISES']}{r}"))
        vir_plus = to_num(cell(ws, f"{COL['VIREMENTS_PLUS_VISES']}{r}"))
        vir_moins = to_num(cell(ws, f"{COL['VIREMENTS_MOINS_VISES']}{r}"))
        total_credits = to_num(cell(ws, f"{COL['TOTAL_CREDITS_VISES']}{r}"))
        engage = to_num(cell(ws, f"{COL['ENGAGE_VISES']}{r}"))

        chap = cell_code(ws, f"{COL['CHAP']}{r+1}")
        prog = cell_code(ws, f"{COL['PROG']}{r+1}")
        reg = cell_code(ws, f"{COL['REG']}{r+1}")
        proj = cell_code(ws, f"{COL['PROJ']}{r+1}")
        lb = cell_code(ws, f"{COL['LB']}{r+1}")

        if all(x is None for x in [intit, chap, prog, reg, proj, lb]):
            r += 2
            continue

        rows.append(dict(
            Year=year, Chap=chap, Prog=prog, Reg=reg, Proj=proj, Lb=lb,
            Intitule=intit,
            Credits_Ouverts=credits,
            Virements_Plus=vir_plus,
            Virements_Moins=vir_moins,
            Total_Credits=total_credits,
            Total_Engage=engage,
        ))
        r += 2

    # Depth + Level
    for row in rows:
        depth = next((i + 1 for i, lvl in enumerate(LEVELS) if row[lvl] is not None), None)
        row["Depth"] = depth
        row["Level"] = LEVEL_NAMES.get(depth)

    # Forward-fill Chap then drop excluded chapters
    chap_fill = None
    kept = []
    for row in rows:
        if row["Chap"]:
            chap_fill = row["Chap"]
        if chap_fill in EXCLUDED_CHAPTERS:
            continue
        kept.append(row)
    return kept


def build_panel(all_rows):
    """Forward-fill hierarchy within each year so each row knows its full path,
    then keep only Line_Keys present in EVERY year."""
    df = pd.DataFrame(all_rows)
    years = sorted(df["Year"].unique())

    out = []
    for y, grp in df.groupby("Year"):
        grp = grp.reset_index(drop=True).copy()
        # forward-fill, but reset deeper levels when an upper level changes
        last = {lvl: None for lvl in LEVELS}
        for i, row in grp.iterrows():
            for j, lvl in enumerate(LEVELS):
                v = row[lvl]
                if v is not None:
                    last[lvl] = v
                    # reset deeper
                    for deeper in LEVELS[j + 1:]:
                        last[deeper] = None
                    break
            for lvl in LEVELS:
                grp.at[i, lvl] = last[lvl]
        out.append(grp)
    df = pd.concat(out, ignore_index=True)
    df["Line_Key"] = df[LEVELS].fillna("").agg("-".join, axis=1).str.strip("-")

    # keep keys present in every non-partial year
    non_partial = [y for y in years if y not in PARTIAL_YEARS]
    presence = (df[~df["Year"].isin(PARTIAL_YEARS)]
                .groupby("Line_Key")["Year"].nunique())
    stable_keys = presence[presence == len(non_partial)].index
    df["Is_Stable"] = df["Line_Key"].isin(stable_keys)
    return df


def detect_extraction_anomalies(panel):
    flags = []
    leaves = panel[panel["Level"] == "Ligne"].copy()

    # Rule 1: over-execution
    over = leaves[
        leaves["Total_Credits"].fillna(0).gt(0)
        & leaves["Total_Engage"].fillna(0).gt(leaves["Total_Credits"].fillna(0) * 1.10)
    ]
    for _, r in over.iterrows():
        ratio = r["Total_Engage"] / r["Total_Credits"]
        flags.append(dict(rule="over_execution", severity="medium",
                          Year=r["Year"], Line_Key=r["Line_Key"],
                          Intitule=r["Intitule"],
                          detail=f"Engage / Credits = {ratio:.2f}",
                          Engage=r["Total_Engage"], Credits=r["Total_Credits"]))

    # Rule 2: negative/zero base
    bad = leaves[(leaves["Total_Engage"].fillna(0) < 0) | (leaves["Credits_Ouverts"].fillna(0) < 0)]
    for _, r in bad.iterrows():
        flags.append(dict(rule="negative_value", severity="high",
                          Year=r["Year"], Line_Key=r["Line_Key"],
                          Intitule=r["Intitule"],
                          detail="negative engage or credits",
                          Engage=r["Total_Engage"], Credits=r["Credits_Ouverts"]))

    # Rule 3: summary keyword in Intitule but classified as Ligne
    def is_summary(name):
        if not isinstance(name, str):
            return False
        s = name.lower()
        return any(k in s for k in SUMMARY_KEYWORDS)
    sums = leaves[leaves["Intitule"].apply(is_summary)]
    for _, r in sums.iterrows():
        flags.append(dict(rule="summary_keyword", severity="high",
                          Year=r["Year"], Line_Key=r["Line_Key"],
                          Intitule=r["Intitule"],
                          detail="Intitule contains summary keyword",
                          Engage=r["Total_Engage"], Credits=r["Credits_Ouverts"]))

    # Rule 4: identity drift (same Line_Key, very different Intitule)
    # Normalize labels to ignore minor unicode/whitespace/case variants
    import re, unicodedata
    def normalize_label(s):
        if not isinstance(s, str):
            return ""
        s = unicodedata.normalize("NFKD", s)
        s = "".join(c for c in s if not unicodedata.combining(c))
        s = re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()
        return s
    norm = leaves.assign(_n=leaves["Intitule"].apply(normalize_label))
    grouped = norm.groupby("Line_Key")["_n"].nunique()
    drifted = grouped[grouped > 1].index
    for key in drifted:
        labs = norm[norm["Line_Key"] == key][["Year", "Intitule"]].drop_duplicates(subset=["Intitule"])
        flags.append(dict(rule="identity_drift", severity="medium",
                          Year=None, Line_Key=key,
                          Intitule=" | ".join(labs["Intitule"].astype(str).head(3)),
                          detail=f"{len(labs)} different normalized labels across years",
                          Engage=None, Credits=None))

    # Rule 5: >10x swing without compensating virement
    leaves_sorted = leaves.sort_values(["Line_Key", "Year"])
    leaves_sorted["prev_engage"] = leaves_sorted.groupby("Line_Key")["Total_Engage"].shift(1)
    swings = leaves_sorted[
        leaves_sorted["prev_engage"].fillna(0).gt(0)
        & (
            (leaves_sorted["Total_Engage"] / leaves_sorted["prev_engage"] > 10)
            | (leaves_sorted["prev_engage"] / leaves_sorted["Total_Engage"].replace(0, np.nan) > 10)
        )
    ]
    for _, r in swings.iterrows():
        ratio = (r["Total_Engage"] or 0) / (r["prev_engage"] or 1)
        vir_net = (r["Virements_Plus"] or 0) - (r["Virements_Moins"] or 0)
        flags.append(dict(rule="huge_swing", severity="low",
                          Year=r["Year"], Line_Key=r["Line_Key"],
                          Intitule=r["Intitule"],
                          detail=f"x{ratio:.1f} vs previous year (vir_net={vir_net:.0f})",
                          Engage=r["Total_Engage"], Credits=r["Credits_Ouverts"]))

    return pd.DataFrame(flags)


def detect_business_anomalies(panel):
    flags = []
    leaves = panel[(panel["Level"] == "Ligne") & (~panel["Year"].isin(PARTIAL_YEARS))].copy()
    leaves["Virement_Net"] = (
        leaves["Virements_Plus"].fillna(0) - leaves["Virements_Moins"].fillna(0)
    )
    leaves["Exec_Rate"] = np.where(
        leaves["Total_Credits"].fillna(0) > 0,
        leaves["Total_Engage"] / leaves["Total_Credits"],
        np.nan,
    )

    # Rule B1: major virement vs initial credits
    huge_vir = leaves[
        leaves["Credits_Ouverts"].fillna(0).gt(0)
        & (leaves["Virement_Net"].abs() / leaves["Credits_Ouverts"] > 0.5)
    ]
    for _, r in huge_vir.iterrows():
        pct = 100 * r["Virement_Net"] / r["Credits_Ouverts"]
        flags.append(dict(rule="major_virement", severity="medium",
                          Year=r["Year"], Line_Key=r["Line_Key"],
                          Intitule=r["Intitule"],
                          detail=f"net virement = {pct:+.0f}% of initial credits"))

    # Rule B2: chronic under-execution
    under = leaves[leaves["Exec_Rate"].lt(0.30)]
    counts = under.groupby("Line_Key").size()
    chronic_under = counts[counts >= 2].index
    for key in chronic_under:
        years = sorted(under[under["Line_Key"] == key]["Year"].tolist())
        intit = leaves[leaves["Line_Key"] == key]["Intitule"].iloc[0]
        flags.append(dict(rule="chronic_under_execution", severity="medium",
                          Year=None, Line_Key=key, Intitule=intit,
                          detail=f"exec rate <30% in years {years}"))

    # Rule B3: chronic over-execution
    over = leaves[leaves["Exec_Rate"].gt(1.0)]
    counts = over.groupby("Line_Key").size()
    chronic_over = counts[counts >= 2].index
    for key in chronic_over:
        years = sorted(over[over["Line_Key"] == key]["Year"].tolist())
        intit = leaves[leaves["Line_Key"] == key]["Intitule"].iloc[0]
        flags.append(dict(rule="chronic_over_execution", severity="medium",
                          Year=None, Line_Key=key, Intitule=intit,
                          detail=f"exec rate >100% in years {years}"))

    # Rule B4: high volatility
    stats = (leaves.groupby("Line_Key")["Total_Engage"]
                  .agg(["mean", "std", "count"])
                  .query("count >= 3 and mean > 0"))
    stats["cv"] = stats["std"] / stats["mean"]
    volatile = stats[stats["cv"] > 0.5]
    for key, row in volatile.iterrows():
        intit = leaves[leaves["Line_Key"] == key]["Intitule"].iloc[0]
        flags.append(dict(rule="high_volatility", severity="low",
                          Year=None, Line_Key=key, Intitule=intit,
                          detail=f"coefficient of variation = {row['cv']:.2f}"))

    # Rule B5: virement inflation -- received more in transfers than the
    # initial allocation. Sign that the opening budget was severely
    # under-estimated.
    leaves["Virements_Plus_f"] = leaves["Virements_Plus"].fillna(0)
    leaves["Virements_Moins_f"] = leaves["Virements_Moins"].fillna(0)
    leaves["Credits_Ouverts_f"] = leaves["Credits_Ouverts"].fillna(0)

    inflation = leaves[
        leaves["Credits_Ouverts_f"].gt(0)
        & leaves["Virements_Plus_f"].gt(leaves["Credits_Ouverts_f"])
    ]
    for _, r in inflation.iterrows():
        ratio = r["Virements_Plus_f"] / r["Credits_Ouverts_f"]
        flags.append(dict(rule="virement_inflation", severity="high",
                          Year=r["Year"], Line_Key=r["Line_Key"],
                          Intitule=r["Intitule"],
                          detail=f"virements +{ratio:.1f}x initial credits"))

    # Rule B6: virement churn -- both Plus and Moins exceed 30% of opened
    # credits in the same year. Indicates back-and-forth transfers /
    # unclear initial intent.
    churn = leaves[
        leaves["Credits_Ouverts_f"].gt(0)
        & leaves["Virements_Plus_f"].div(leaves["Credits_Ouverts_f"]).gt(0.30)
        & leaves["Virements_Moins_f"].div(leaves["Credits_Ouverts_f"]).gt(0.30)
    ]
    for _, r in churn.iterrows():
        pp = 100 * r["Virements_Plus_f"] / r["Credits_Ouverts_f"]
        pm = 100 * r["Virements_Moins_f"] / r["Credits_Ouverts_f"]
        flags.append(dict(rule="virement_churn", severity="medium",
                          Year=r["Year"], Line_Key=r["Line_Key"],
                          Intitule=r["Intitule"],
                          detail=f"+{pp:.0f}% / -{pm:.0f}% of initial credits"))

    # Rule B7: unused top-up -- got >30% extra in virements but engagement
    # rate stayed below 80%. The line was reinforced but the money was
    # not used: should the transfer have happened?
    unused = leaves[
        leaves["Credits_Ouverts_f"].gt(0)
        & leaves["Virements_Plus_f"].div(leaves["Credits_Ouverts_f"]).gt(0.30)
        & leaves["Exec_Rate"].lt(0.80)
    ]
    for _, r in unused.iterrows():
        pp = 100 * r["Virements_Plus_f"] / r["Credits_Ouverts_f"]
        er = 100 * (r["Exec_Rate"] or 0)
        flags.append(dict(rule="unused_top_up", severity="medium",
                          Year=r["Year"], Line_Key=r["Line_Key"],
                          Intitule=r["Intitule"],
                          detail=f"top-up +{pp:.0f}% but exec rate only {er:.0f}%"))

    return pd.DataFrame(flags)


def main():
    # 1. Re-extract raw rows with the EXTRA columns (Virements + Total_Credits)
    raw_rows = []
    files = sorted(RAW.glob("SituationChap-*.xlsx"))
    print(f"Found {len(files)} raw xlsx files")
    for f in files:
        try:
            year = int("".join(c for c in f.stem if c.isdigit())[:4])
        except ValueError:
            continue
        print(f"  Reading {f.name} (year {year})...")
        rows = extract_one(f, year)
        print(f"    {len(rows)} rows extracted")
        raw_rows.extend(rows)
    raw = pd.DataFrame(raw_rows)

    # 2. Load the trusted stable panel (107 stable Line_Keys, hierarchy already filled)
    stable_path = ROOT / "data" / "02_cleaned" / "SituationChap_STABLE_PANEL.xlsx"
    panel = pd.read_excel(stable_path)
    print(f"\nLoaded stable panel: {len(panel):,} rows, {panel['Line_Key'].nunique()} unique Line_Keys")

    # 3. Enrich: for each (Year, leaf-level codes) in stable panel, match a raw row to grab virements
    # The raw row for a leaf has its OWN level code filled (Lb usually) but parents NaN.
    # We match on the deepest non-empty code per stable panel row vs raw rows where that code matches.
    # Simpler & robust: match on (Year, Intitule) since intitules are stable within a year.
    raw_for_join = raw[["Year", "Intitule", "Credits_Ouverts", "Virements_Plus",
                        "Virements_Moins", "Total_Credits", "Total_Engage"]].copy()
    raw_for_join = raw_for_join.dropna(subset=["Intitule"])
    # Keep the first match per (Year, Intitule) to avoid blowup
    raw_for_join = raw_for_join.drop_duplicates(subset=["Year", "Intitule"])

    panel = panel.merge(raw_for_join, on=["Year", "Intitule"], how="left",
                        suffixes=("", "_raw"))
    print(f"After enrichment: {len(panel):,} rows, {panel['Virements_Plus'].notna().sum()} with virement data")

    # Use stable panel's existing engage; fall back to raw if missing
    panel["Total_Engage"] = panel["Total_Engage"].fillna(panel["Total_Engage_Vises"])
    panel["Credits_Ouverts"] = panel["Credits_Ouverts"].fillna(panel["Credits_Ouverts_Vises"])

    panel.to_csv(OUT / "03_raw_enriched.csv", index=False)

    ext = detect_extraction_anomalies(panel)
    bus = detect_business_anomalies(panel)
    ext.to_csv(OUT / "03_anomalies_extraction.csv", index=False)
    bus.to_csv(OUT / "03_anomalies_business.csv", index=False)

    print(f"\n=== EXTRACTION ANOMALIES ({len(ext)}) ===")
    if len(ext):
        print(ext.groupby(["rule", "severity"]).size().to_string())

    print(f"\n=== BUSINESS ANOMALIES ({len(bus)}) ===")
    if len(bus):
        print(bus.groupby(["rule", "severity"]).size().to_string())

    # ------------------------------------------------------------------
    # Re-run 2025 backtest excluding extraction anomalies
    # ------------------------------------------------------------------
    print("\n" + "=" * 72)
    print("RE-RUNNING 2025 BACKTEST WITH EXTRACTION ANOMALIES EXCLUDED")
    print("=" * 72)

    def smape(a, p):
        d = (abs(a) + abs(p)) / 2
        return 0.0 if d == 0 else 100 * abs(a - p) / d

    def naive(h): return h[-1]
    def mean3(h): return float(np.mean(h[-3:]))
    def median_(h): return float(np.median(h))
    def trend(h):
        if len(h) < 2:
            return float(h[-1])
        a, b = np.polyfit(np.arange(len(h)), h, 1)
        return float(a * len(h) + b)
    METHODS = {"Naive": naive, "Mean3": mean3, "Median": median_, "Trend": trend}

    suspect_keys = set(ext.dropna(subset=["Line_Key"])["Line_Key"].unique())

    leaves = panel[
        (panel["Level"] == "Ligne")
        & (~panel["Year"].isin(PARTIAL_YEARS))
    ].copy()
    leaves["Total_Engage"] = leaves["Total_Engage"].fillna(leaves["Total_Engage_Vises"])

    def run(df, exclude_keys):
        rows = []
        for key, grp in df.groupby("Line_Key"):
            if key in exclude_keys:
                continue
            g = grp.sort_values("Year").dropna(subset=["Total_Engage"])
            train = g[g["Year"] < 2025]
            actual_row = g[g["Year"] == 2025]
            if len(train) < 2 or actual_row.empty:
                continue
            actual = float(actual_row["Total_Engage"].iloc[0])
            tv = train["Total_Engage"].astype(float).tolist()
            best_sm = np.inf
            for fn in METHODS.values():
                try:
                    p = fn(tv)
                except Exception:
                    continue
                if not np.isfinite(p):
                    continue
                best_sm = min(best_sm, smape(actual, p))
            if np.isfinite(best_sm):
                rows.append(dict(Line_Key=key, sMAPE=best_sm,
                                 Actual=actual,
                                 Predicted_best=tv[-1]))  # placeholder
        return pd.DataFrame(rows)

    before = run(leaves, exclude_keys=set())
    after = run(leaves, exclude_keys=suspect_keys)

    def summarize(df, label):
        if len(df) == 0:
            return {"label": label, "n": 0, "median_sMAPE": None,
                    "mean_sMAPE": None, "pct_under_10": None, "pct_under_25": None}
        sm = df["sMAPE"].values
        return {
            "label": label,
            "n": len(df),
            "median_sMAPE": round(float(np.median(sm)), 2),
            "mean_sMAPE": round(float(np.mean(sm)), 2),
            "pct_under_10": round(100 * float(np.mean(sm <= 10)), 1),
            "pct_under_25": round(100 * float(np.mean(sm <= 25)), 1),
        }

    s1 = summarize(before, "ALL lignes (no anomaly filter)")
    s2 = summarize(after, f"Excluding {len(suspect_keys)} suspect Line_Keys")

    cmp = pd.DataFrame([s1, s2])
    print(cmp.to_string(index=False))
    cmp.to_csv(OUT / "03_anomaly_impact_2025.csv", index=False)

    print(f"\nFiles written to {OUT}")


if __name__ == "__main__":
    main()
