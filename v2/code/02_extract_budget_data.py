"""Extract clean budget data from raw SituationChap-YYYY.xlsx files.

Raw layout: each budget line spans 2 physical rows
  - top row    : Intitule (col H), Credits Ouverts Vises (K), Total Engage Vises (AH)
  - bottom row : Chap (B), Prog (C), Reg (D), Proj (E), Lb (G)

We keep only the two "Vises" amounts requested by the manager.
Output: one cleaned .xlsx per year in v2/data/cleaned/.
"""
from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
RAW_DIR = DATA_DIR / "01_raw"
OUT_DIR = DATA_DIR / "02_cleaned"
OUT_DIR.mkdir(exist_ok=True)

# Column letters in the raw file
COL_CHAP, COL_PROG, COL_REG, COL_PROJ, COL_LB = "B", "C", "D", "E", "G"
COL_INTITULE = "H"
COL_CREDITS_OUVERTS_VISES = "K"
COL_TOTAL_ENGAGE_VISES = "AH"

DATA_START_ROW = 6  # row 4-5 are headers


def cell(ws, col_letter: str, row: int):
    v = ws[f"{col_letter}{row}"].value
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def cell_code(ws, col_letter: str, row: int):
    """Read hierarchy code as string, preserving leading zeros."""
    v = ws[f"{col_letter}{row}"].value
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def extract_year(file_path: Path) -> int | None:
    m = re.search(r"(\d{4})", file_path.stem)
    return int(m.group(1)) if m else None


def extract_one(xlsx_path: Path) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    year = extract_year(xlsx_path)

    rows = []
    r = DATA_START_ROW
    while r <= ws.max_row:
        intitule = cell(ws, COL_INTITULE, r)
        credits = cell(ws, COL_CREDITS_OUVERTS_VISES, r)
        engage = cell(ws, COL_TOTAL_ENGAGE_VISES, r)

        chap = cell_code(ws, COL_CHAP, r + 1)
        prog = cell_code(ws, COL_PROG, r + 1)
        reg = cell_code(ws, COL_REG, r + 1)
        proj = cell_code(ws, COL_PROJ, r + 1)
        lb = cell_code(ws, COL_LB, r + 1)

        if intitule is None and chap is None and prog is None and reg is None and proj is None and lb is None:
            r += 2
            continue  # blank pair

        rows.append({
            "Year": year,
            "Chap": chap,
            "Prog": prog,
            "Reg": reg,
            "Proj": proj,
            "Lb": lb,
            "Intitule": intitule,
            "Credits_Ouverts_Vises": credits,
            "Total_Engage_Vises": engage,
        })
        r += 2

    df = pd.DataFrame(rows)
    # Force numeric for the two amounts
    for c in ("Credits_Ouverts_Vises", "Total_Engage_Vises"):
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # Depth = how deep this row sits in the budget tree.
    # 1=Chapitre, 2=Programme, 3=Region, 4=Projet, 5=Ligne (leaf).
    # Each raw row has exactly ONE code column filled; depth = which one.
    LEVELS = ["Chap", "Prog", "Reg", "Proj", "Lb"]
    LEVEL_NAMES = {1: "Chapitre", 2: "Programme", 3: "Region", 4: "Projet", 5: "Ligne"}

    def _row_depth(r):
        for i, col in enumerate(LEVELS, start=1):
            if pd.notna(r[col]):
                return i
        return pd.NA

    df["Depth"] = df[LEVELS].apply(_row_depth, axis=1)
    df["Level"] = df["Depth"].map(LEVEL_NAMES)

    # Drop chapters the manager said are not relevant.
    # 3200106002 = Fonds d'entraide familiale
    # 3200106001 = Fonds spécial pour le soutien des juridictions
    EXCLUDED_CHAPTERS = {"3200106002", "3200106001"}
    chap_filled = df["Chap"].ffill()
    df = df[~chap_filled.isin(EXCLUDED_CHAPTERS)].reset_index(drop=True)

    # Reorder so Depth/Level sit right after the code columns
    cols = ["Year", "Chap", "Prog", "Reg", "Proj", "Lb", "Depth", "Level",
            "Intitule", "Credits_Ouverts_Vises", "Total_Engage_Vises"]
    return df[cols]


def main():
    files = sorted(p for p in RAW_DIR.glob("SituationChap-*.xlsx") if re.fullmatch(r"SituationChap-\d{4}", p.stem))
    if not files:
        raise SystemExit(f"No yearly SituationChap-YYYY.xlsx in {RAW_DIR}")

    all_dfs = []
    for f in files:
        print(f"Processing {f.name}")
        df = extract_one(f)
        out = OUT_DIR / f"{f.stem}_clean.xlsx"
        df.to_excel(out, index=False)
        print(f"  {len(df)} rows -> {out.relative_to(DATA_DIR.parent.parent)}")
        all_dfs.append(df)

    combined = pd.concat(all_dfs, ignore_index=True)
    combined_path = OUT_DIR / "SituationChap_ALL_YEARS.xlsx"
    combined.to_excel(combined_path, index=False)
    print(f"\nCombined: {len(combined)} rows -> {combined_path.relative_to(DATA_DIR.parent.parent)}")

    stable = build_stable_panel(combined)
    stable_path = OUT_DIR / "SituationChap_STABLE_PANEL.xlsx"
    stable.to_excel(stable_path, index=False)
    n_keys = stable["Line_Key"].nunique()
    n_years = stable["Year"].nunique()
    print(f"Stable panel: {len(stable)} rows ({n_keys} lines x {n_years} years) -> {stable_path.relative_to(DATA_DIR.parent.parent)}")


def build_stable_panel(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only budget lines present in EVERY year.

    Each row in the raw data sits at a single hierarchy level (Chap or Prog or
    Reg or Proj or Lb). We forward-fill the parent codes within each year so
    every row carries its full path, then build a canonical key per row.
    """
    df = df.copy()
    levels = ["Chap", "Prog", "Reg", "Proj", "Lb"]

    # Determine each row's depth = index of the deepest non-null level
    depth = df[levels].notna().to_numpy().cumsum(axis=1).argmax(axis=1)
    df["_Depth"] = depth

    # Forward-fill parent codes within each year so each row knows its full path
    for col in levels:
        df[f"{col}_ff"] = df.groupby("Year")[col].ffill()

    # Mask out levels deeper than the row's own depth (those parents only)
    ff_cols = [f"{c}_ff" for c in levels]
    for i, col in enumerate(ff_cols):
        df.loc[df["_Depth"] < i, col] = pd.NA

    # Canonical key = full path up to the row's level
    df["Line_Key"] = df[ff_cols].astype("string").agg(
        lambda r: "-".join(x for x in r if pd.notna(x)), axis=1
    )

    # Keep keys present in all years
    n_years = df["Year"].nunique()
    counts = df.groupby("Line_Key")["Year"].nunique()
    stable_keys = counts[counts == n_years].index
    stable = df[df["Line_Key"].isin(stable_keys)].copy()

    # Recompute Depth/Level from Line_Key (it encodes the full path)
    LEVEL_NAMES = {1: "Chapitre", 2: "Programme", 3: "Region", 4: "Projet", 5: "Ligne"}
    stable["Depth"] = stable["Line_Key"].str.count("-") + 1
    stable["Level"] = stable["Depth"].map(LEVEL_NAMES)

    keep_cols = ["Year", "Line_Key", *levels, "Depth", "Level", "Intitule",
                 "Credits_Ouverts_Vises", "Total_Engage_Vises"]
    return stable[keep_cols].sort_values(["Line_Key", "Year"]).reset_index(drop=True)


if __name__ == "__main__":
    main()
