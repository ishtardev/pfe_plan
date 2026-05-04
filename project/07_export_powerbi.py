"""
07_export_powerbi.py
--------------------
Produit un fichier Excel propre, format\u00e9 et pr\u00eat \u00e0 \u00eatre import\u00e9 dans Power BI
(ou pr\u00e9sent\u00e9 tel quel \u00e0 un encadrant non technique).

Entr\u00e9e : project/previsions_2025_20260429.xlsx
Sortie : project/previsions_2025_powerbi.xlsx

Trois onglets :
    1. Pr\u00e9visions       - table principale, format\u00e9e, couleurs sur Risque
    2. Synth\u00e8se          - KPI + agr\u00e9gats par cat\u00e9gorie / r\u00e9gion / programme
    3. Anomalies         - lignes flagg\u00e9es uniquement
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
SRC = ROOT / "previsions_2025_20260429.xlsx"
HIST_AGG = ROOT / "data_long.csv"
HIST_LINES = ROOT / "data_lines.csv"
DST = ROOT / "previsions_2025_powerbi.xlsx"

# ---------------------------------------------------------------- 1. Charger
df = pd.read_excel(SRC)
hist_agg = pd.read_csv(HIST_AGG)
hist_lines = pd.read_csv(HIST_LINES)

# Tri par MDH \u00e0 risque d\u00e9croissant pour que les lignes les plus importantes
# apparaissent en haut.
df = df.sort_values("MDH a risque", ascending=False).reset_index(drop=True)

# ---------------------------------------------------------------- 2. Synth\u00e8se
total_lignes = len(df)
total_lf = df["LF (MDH)"].sum()
total_risque = df["MDH a risque"].sum()
nb_alerte = (df["Risque"] == "Alerte").sum()
nb_attention = (df["Risque"] == "Attention").sum()
nb_anomalies = df["Anomalie"].fillna("").astype(str).str.strip().ne("").sum()
taux_moyen = df["T4 predit T3 (%)"].mean()

kpi = pd.DataFrame(
    {
        "Indicateur": [
            "Nombre de lignes budg\u00e9taires",
            "Budget total inscrit (MDH)",
            "Taux d'ex\u00e9cution moyen pr\u00e9dit (%)",
            "Lignes en alerte",
            "Lignes \u00e0 surveiller",
            "Anomalies d\u00e9tect\u00e9es",
            "Montant \u00e0 risque (MDH)",
        ],
        "Valeur": [
            total_lignes,
            round(total_lf, 1),
            round(taux_moyen, 1),
            int(nb_alerte),
            int(nb_attention),
            int(nb_anomalies),
            round(total_risque, 1),
        ],
    }
)

agg_cat = (
    df.groupby("Type budget")
    .agg(
        Lignes=("Ligne budgetaire", "count"),
        Budget_MDH=("LF (MDH)", "sum"),
        Taux_moyen_pct=("T4 predit T3 (%)", "mean"),
        Risque_MDH=("MDH a risque", "sum"),
    )
    .round(1)
    .reset_index()
)

agg_prog = (
    df.groupby("Programme")
    .agg(
        Lignes=("Ligne budgetaire", "count"),
        Budget_MDH=("LF (MDH)", "sum"),
        Taux_moyen_pct=("T4 predit T3 (%)", "mean"),
        Risque_MDH=("MDH a risque", "sum"),
    )
    .round(1)
    .sort_values("Risque_MDH", ascending=False)
    .reset_index()
)

agg_reg = (
    df.groupby("Region")
    .agg(
        Lignes=("Ligne budgetaire", "count"),
        Budget_MDH=("LF (MDH)", "sum"),
        Taux_moyen_pct=("T4 predit T3 (%)", "mean"),
        Risque_MDH=("MDH a risque", "sum"),
    )
    .round(1)
    .sort_values("Risque_MDH", ascending=False)
    .reset_index()
)

# ---------------------------------------------------------------- 3. Anomalies
anomalies = df[df["Anomalie"].fillna("").astype(str).str.strip().ne("")].copy()

# ---------------------------------------------------------------- 3b. Historique cat\u00e9gorie
# Format long, pr\u00eat \u00e0 \u00eatre tra\u00e7\u00e9 dans Power BI (line chart).
hist_cat = hist_agg.rename(
    columns={
        "year": "Annee",
        "quarter": "Trimestre",
        "category": "Type budget",
        "lf": "LF (MDH)",
        "realise": "Realise (MDH)",
        "taux": "Taux execution",
    }
)[["Annee", "Trimestre", "Type budget", "LF (MDH)", "Realise (MDH)", "Taux execution"]]
hist_cat["Taux execution (%)"] = (hist_cat["Taux execution"] * 100).round(1)
hist_cat = hist_cat.drop(columns=["Taux execution"])

# ---------------------------------------------------------------- 3c. Pr\u00e9dit vs R\u00e9el (par cat\u00e9gorie, T4)
# Pour chaque ann\u00e9e, on extrait le taux T4 r\u00e9el et on le compare au taux moyen
# pr\u00e9dit pour 2025 (issu de la table de pr\u00e9visions, agr\u00e9g\u00e9 pond\u00e9r\u00e9 par LF).
real_t4 = (
    hist_agg[hist_agg["quarter"] == "T4"]
    .assign(taux_pct=lambda x: (x["taux"] * 100).round(1))[
        ["year", "category", "taux_pct"]
    ]
    .rename(columns={"year": "Annee", "category": "Type budget", "taux_pct": "T4 reel (%)"})
)

# Pr\u00e9diction 2025 agr\u00e9g\u00e9e par cat\u00e9gorie (moyenne pond\u00e9r\u00e9e par LF)
def _wavg(group, col, w="LF (MDH)"):
    return (group[col] * group[w]).sum() / group[w].sum()

pred_2025 = (
    df.groupby("Type budget")
    .apply(lambda g: pd.Series({"T4 predit (%)": round(_wavg(g, "T4 predit T3 (%)"), 1)}))
    .reset_index()
)
pred_2025["Annee"] = 2025

pred_vs_real = real_t4.merge(pred_2025, on=["Annee", "Type budget"], how="outer")
pred_vs_real["Ecart (pp)"] = (
    pred_vs_real["T4 reel (%)"] - pred_vs_real["T4 predit (%)"]
).round(1)
pred_vs_real = pred_vs_real.sort_values(["Type budget", "Annee"]).reset_index(drop=True)

# ---------------------------------------------------------------- 3d. Historique par ligne
# Pour les 130 lignes de la table Pr\u00e9visions, leur taux T4 historique 2020-2024.
hist_l = hist_lines.copy()
hist_l = hist_l.rename(
    columns={
        "year": "Annee",
        "quarter": "Trimestre",
        "ligne_label": "Ligne budgetaire",
        "region_label": "Region",
        "programme_label": "Programme",
        "budget_type": "Type budget",
        "lf_mdh": "LF (MDH)",
        "taux": "Taux",
    }
)
hist_l["Taux (%)"] = (hist_l["Taux"] * 100).round(1)
hist_l = hist_l[
    [
        "Annee",
        "Trimestre",
        "Programme",
        "Region",
        "Ligne budgetaire",
        "Type budget",
        "LF (MDH)",
        "Taux (%)",
    ]
]

# ---------------------------------------------------------------- 4. \u00c9crire
with pd.ExcelWriter(DST, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Pr\u00e9visions", index=False)

    # Synth\u00e8se : on empile les blocs avec des titres
    start = 0
    kpi.to_excel(writer, sheet_name="Synth\u00e8se", index=False, startrow=start)
    start += len(kpi) + 3
    pd.DataFrame({"": ["Par cat\u00e9gorie"]}).to_excel(
        writer, sheet_name="Synth\u00e8se", index=False, header=False, startrow=start
    )
    start += 1
    agg_cat.to_excel(writer, sheet_name="Synth\u00e8se", index=False, startrow=start)
    start += len(agg_cat) + 3
    pd.DataFrame({"": ["Par programme"]}).to_excel(
        writer, sheet_name="Synth\u00e8se", index=False, header=False, startrow=start
    )
    start += 1
    agg_prog.to_excel(writer, sheet_name="Synth\u00e8se", index=False, startrow=start)
    start += len(agg_prog) + 3
    pd.DataFrame({"": ["Par r\u00e9gion"]}).to_excel(
        writer, sheet_name="Synth\u00e8se", index=False, header=False, startrow=start
    )
    start += 1
    agg_reg.to_excel(writer, sheet_name="Synth\u00e8se", index=False, startrow=start)

    anomalies.to_excel(writer, sheet_name="Anomalies", index=False)
    hist_cat.to_excel(writer, sheet_name="Historique cat\u00e9gorie", index=False)
    pred_vs_real.to_excel(writer, sheet_name="Pr\u00e9dit vs R\u00e9el", index=False)
    hist_l.to_excel(writer, sheet_name="Historique lignes", index=False)

# ---------------------------------------------------------------- 5. Mise en forme
wb = load_workbook(DST)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # bleu fonc\u00e9
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12, color="1F3864")
ALERTE_FILL = PatternFill("solid", fgColor="F8CBAD")
ATTENTION_FILL = PatternFill("solid", fgColor="FFE699")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws, max_width=42):
    for col_idx, col_cells in enumerate(ws.columns, 1):
        values = [str(c.value) if c.value is not None else "" for c in col_cells]
        width = min(max(len(v) for v in values) + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)


# ----- Onglet Pr\u00e9visions
ws = wb["Pr\u00e9visions"]
style_header(ws)
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
autosize(ws)

# Coloration conditionnelle sur la colonne Risque
risque_col = None
for idx, cell in enumerate(ws[1], 1):
    if cell.value == "Risque":
        risque_col = get_column_letter(idx)
        break

if risque_col:
    rng = f"{risque_col}2:{risque_col}{ws.max_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Alerte"'], fill=ALERTE_FILL)
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Attention"'], fill=ATTENTION_FILL),
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"OK"'], fill=OK_FILL)
    )

# Format pourcentage / d\u00e9cimal
for col_name, fmt in {
    "T2 realise (%)": "0.0",
    "T4 predit T2 (%)": "0.0",
    "T4 predit T3 (%)": "0.0",
    "T4 reel (%)": "0.0",
    "LF (MDH)": "#,##0.00",
    "MDH a risque": "#,##0.00",
    "Ecart T3 (pp)": "+0.0;-0.0;0.0",
}.items():
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == col_name:
            letter = get_column_letter(idx)
            for row in range(2, ws.max_row + 1):
                ws[f"{letter}{row}"].number_format = fmt
            break

# ----- Onglet Synth\u00e8se : juste embellir un peu
ws = wb["Synth\u00e8se"]
ws.column_dimensions["A"].width = 36
for col in ["B", "C", "D", "E"]:
    ws.column_dimensions[col].width = 18

# Mettre en gras les lignes de titre de section et les ent\u00eates de tableau
for row in ws.iter_rows():
    first = row[0].value
    if first in {"Indicateur", "Type budget", "Programme", "Region"}:
        for cell in row:
            if cell.value is not None:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
    elif first in {"Par cat\u00e9gorie", "Par programme", "Par r\u00e9gion"}:
        row[0].font = SECTION_FONT

# ----- Onglet Anomalies
ws = wb["Anomalies"]
style_header(ws)
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
autosize(ws)

# ----- Onglets historiques
for sheet_name in ["Historique cat\u00e9gorie", "Pr\u00e9dit vs R\u00e9el", "Historique lignes"]:
    ws = wb[sheet_name]
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)

wb.save(DST)
print(f"OK -> {DST}")
print(f"  {total_lignes} lignes, {nb_alerte} en alerte, {nb_attention} \u00e0 surveiller, "
      f"{nb_anomalies} anomalies, {total_risque:.1f} MDH \u00e0 risque")
