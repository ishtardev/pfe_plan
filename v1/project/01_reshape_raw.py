"""
Reshapes raw_data.xlsx (wide messy format) into a clean long-format CSV
ready for ML modeling.

Output columns:
    year        : budget year (2020-2025)
    quarter     : T1, T2, T3, T4
    quarter_num : 1, 2, 3, 4
    category    : PERSONNEL | MATERIEL | INVESTISSEMENT
    lf          : credit alloue (LF initial, MDH)
    realise     : montant paye cumule au trimestre (MDH)
    taux        : realise / lf (taux d'execution cumule)

Then adds ML-ready features:
    taux_lag1   : taux du trimestre precedent (meme annee)
    taux_lag4   : taux du meme trimestre, annee precedente
    lf_ratio    : lf / moyenne LF de la categorie (normalisation)
"""

import openpyxl
import pandas as pd

# --- 1. Parse raw file ---------------------------------------------------

wb = openpyxl.load_workbook("raw_data.xlsx")
ws = wb["Feuil1"]

rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

# Column offsets where each year block starts (identified from structure)
YEAR_COL_STARTS = [0, 5, 10, 16, 22, 28]

records = []

i = 0
while i < len(rows):
    row = rows[i]

    # Detect a quarter header row (first cell is T1/T2/T3/T4)
    if row[0] in ("T1", "T2", "T3", "T4"):
        quarter = row[0]

        # Extract year from the date field (col+2) in each year block
        year_map = {}
        for col_start in YEAR_COL_STARTS:
            date_val = row[col_start + 2]
            if date_val is not None and hasattr(date_val, "year"):
                year_map[col_start] = date_val.year

        # Next 3 rows are the category data rows
        for j in range(1, 4):
            if i + j >= len(rows):
                break
            data_row = rows[i + j]
            if data_row[0] is None:
                break

            for col_start, year in year_map.items():
                raw_cat = str(data_row[col_start + 0]).strip()
                lf_val  = data_row[col_start + 1]
                real_val = data_row[col_start + 2]
                taux_val = data_row[col_start + 3]

                # Normalize category label
                if "PERSONNEL" in raw_cat.upper():
                    cat = "PERSONNEL"
                elif "INVESTISSEMENT" in raw_cat.upper():
                    cat = "INVESTISSEMENT"
                elif "MATERIEL" in raw_cat.upper():
                    cat = "MATERIEL"
                else:
                    cat = raw_cat

                # Clean numeric values (some are stored as strings e.g. "1 157,2")
                def parse_num(v):
                    if v is None:
                        return None
                    if isinstance(v, (int, float)):
                        return float(v)
                    v = str(v).replace("\xa0", "").replace(" ", "").replace(",", ".")
                    try:
                        return float(v)
                    except ValueError:
                        return None

                records.append({
                    "year":     year,
                    "quarter":  quarter,
                    "quarter_num": int(quarter[1]),
                    "category": cat,
                    "lf":       parse_num(lf_val),
                    "realise":  parse_num(real_val),
                    "taux":     parse_num(taux_val),
                })

        i += 4  # skip header + 3 data rows
    else:
        i += 1

# --- 2. Build clean dataframe --------------------------------------------

df = pd.DataFrame(records).drop_duplicates()
df = df.sort_values(["category", "year", "quarter_num"]).reset_index(drop=True)

# --- 3. Add ML features --------------------------------------------------

# taux_lag1 : taux of previous quarter (same year)
df["taux_lag1"] = df.groupby(["category", "year"])["taux"].shift(1)

# taux_lag4 : taux of same quarter, previous year
df["taux_lag4"] = df.groupby(["category", "quarter"])["taux"].shift(1)

# lf_ratio : LF relative to the category's average LF across all years
mean_lf = df.groupby("category")["lf"].transform("mean")
df["lf_ratio"] = df["lf"] / mean_lf

# --- 4. Save outputs -----------------------------------------------------

df.to_csv("data_long.csv", index=False, sep=";")

print("Saved data_long.csv")
print(f"\nShape: {df.shape}")
print(f"\nSample:\n{df.to_string()}")

# Summary stats per category
print("\n--- Taux d'execution year-end (T4) per category ---")
t4 = df[df["quarter"] == "T4"].copy()
print(t4.groupby("category")[["year","lf","realise","taux"]].apply(lambda x: x.to_string(index=False)))
